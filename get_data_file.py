from openpyxl import load_workbook


work_file = load_workbook("db/price.xlsx")
get_first_sheet = work_file.get_sheet_names()[0]
sheet_price = work_file.get_sheet_by_name(str(get_first_sheet))

#print(sheet_price.cell(row=1, column=2).value)
print(sheet_price['A1'].value)