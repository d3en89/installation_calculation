from openpyxl import load_workbook
from core.settings import DATABASES, MEDIA_ROOT
import sqlite3


def main(url, file):
    """ Getting data from xlsx file """
    #work_file = load_workbook("calc/static/tmp/price.xlsx")
    #work_file = load_workbook(f'{url}\\{file}') ## for windows
    work_file = load_workbook(f'{url}/{file}') ## for linux
    get_first_sheet = work_file.sheetnames[0]
    sheet_price = work_file[str(get_first_sheet)]
    data_sheet_price = list(sheet_price.values)[1:]


    """ Check data list 
        if there None in the  first 6 collumns, this row deleted
        """
    count = 1
    while count < len(data_sheet_price):
        if None in data_sheet_price[count][:6]:
            data_sheet_price.pop(count)
            count -= 1
        count += 1


    try:
        """ Module connects and writes data in the sql datasbase
        ********* now he only insert and update rows
        """

        connect_to_db = sqlite3.connect(DATABASES['default']['NAME'])
        cursor = connect_to_db.cursor()

        cursor.execute('select index_sheet, name_of_works, quantity, uom, cost, curency from calc_data_inner')
        answer = cursor.fetchall()
        price_update_db = []

        # Compare value in base and value in sheet
        # if data diverges,  value in sheet add to execute
        for index_base, value_base in enumerate(answer):
            for index_file, value_file in enumerate(data_sheet_price):
                if value_base[0] == value_file[0]:
                    for val in range(len(value_base)):
                        if value_base[val] != value_file[val]:
                            price_update_db.append((value_file[1],value_file[2],value_file[3],value_file[4],value_file[5], value_file[0]))
                    data_sheet_price.pop(index_file)

        if len(data_sheet_price) > 0 :
           sql_insert = 'INSERT INTO calc_data_inner values(NULL, ?, ?, ?, ?, ?, ?)'
           cursor.executemany(sql_insert, data_sheet_price)

        if len(price_update_db) > 0:
            sql_update = """UPDATE calc_data_inner  set name_of_works = ? , quantity = ? , uom = ? , cost = ? , curency = ? where index_sheet LIKE ? """
            cursor.executemany(sql_update, price_update_db)

        connect_to_db.commit()
        cursor.close()

    except sqlite3.DatabaseError as error:
        print("Ошибка при подключении к sqlite data", error)
    except sqlite3.Error as error:
        print("Ошибка при подключении к sqlite", error)
    finally:
        if (connect_to_db):
            connect_to_db.close()